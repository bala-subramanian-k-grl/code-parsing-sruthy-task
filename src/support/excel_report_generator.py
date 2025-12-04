"""
Excel validation report generator (OOP + Polymorphism + Overloading).
"""

from __future__ import annotations

from abc import ABC
from collections.abc import Callable
from pathlib import Path

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore

from src.core.config.models import ParserResult
from src.support.base_report_generator import BaseReportGenerator


class ExcelReportGenerator(BaseReportGenerator, ABC):
    """Generate Excel validation report."""

    # ---------------------------------------------------------
    # Polymorphic Identification
    # ---------------------------------------------------------
    @property
    def report_type(self) -> str:
        """Method implementation."""
        return "Excel"

    @property
    def output_extension(self) -> str:
        """Compatibility alias."""
        return ".xlsx"

    def get_file_extension(self) -> str:
        """Polymorphic format extension."""
        return ".xlsx"     # FIXED ✔ (must return .xlsx, not xlsx)

    # ---------------------------------------------------------
    # Encapsulated Metrics
    # ---------------------------------------------------------
    _METRICS: list[tuple[str, str | Callable[[ParserResult], int]]] = [
        ("Metric", "Value"),
        ("TOC Entries", lambda r: len(r.toc_entries)),
        ("Content Items", lambda r: len(r.content_items)),
    ]

    # ---------------------------------------------------------
    # VALIDATION (from Template Pattern)
    # ---------------------------------------------------------
    def _validate_result(self, result: ParserResult) -> None:
        """Method implementation."""
        if not result.toc_entries and not result.content_items:
            raise ValueError("ParserResult contains no data for Excel report.")

    # ---------------------------------------------------------
    # FORMAT STEP (Template Pattern)
    # ---------------------------------------------------------
    def _format_data(self, result: ParserResult) -> Workbook:
        """Method implementation."""
        workbook: Workbook = Workbook()
        sheet: Worksheet = workbook.active  # type: ignore
        sheet.title = "Validation"

        # Write table rows
        for idx, (metric_name, value_fn) in enumerate(self._METRICS, start=1):
            sheet.cell(row=idx, column=1, value=metric_name)

            value = value_fn(result) if callable(value_fn) else value_fn
            sheet.cell(row=idx, column=2, value=value)

        return workbook

    # ---------------------------------------------------------
    # WRITE STEP (Template Pattern)
    # Returns bytes written
    # ---------------------------------------------------------
    def _write_to_file(self, data: Workbook, path: Path) -> int:
        """
        Save workbook to Excel file.
        Must return number of bytes written for BaseReportGenerator.
        """
        try:
            data.save(path)
            return path.stat().st_size if path.exists() else 0
        except OSError as e:
            msg = f"Failed to save Excel report to '{path}': {e}"
            raise OSError(msg) from e

    # ---------------------------------------------------------
    # OPTIONAL HOOKS
    # ---------------------------------------------------------
    def before_write(self, result: ParserResult, path: Path) -> None:
        """Optional polymorphic hook."""
        # Can add logs, timing, etc.

    def after_write(self, result: ParserResult, path: Path) -> None:
        """Optional polymorphic hook."""

    # ---------------------------------------------------------
    # Magic Methods
    # ---------------------------------------------------------
    def __str__(self) -> str:
        """Method implementation."""
        return f"ExcelReportGenerator(ext={self.get_file_extension()})"

    def __repr__(self) -> str:
        """Method implementation."""
        return "ExcelReportGenerator()"

    def __eq__(self, other: object) -> bool:
        """Method implementation."""
        return isinstance(other, ExcelReportGenerator)

    def __hash__(self) -> int:
        """Method implementation."""
        return hash(self.__class__.__name__)

    def __bool__(self) -> bool:
        """Method implementation."""
        return True

    def __lt__(self, other: object) -> bool:
        """Method implementation."""
        if not isinstance(other, ExcelReportGenerator):
            return NotImplemented
        return len(self) < len(other)

    def __le__(self, other: object) -> bool:
        """Method implementation."""
        return self == other or self < other

    def __int__(self) -> int:
        """Method implementation."""
        return len(self._METRICS)

    def __float__(self) -> float:
        """Method implementation."""
        return float(len(self._METRICS))

    def __len__(self) -> int:
        """Number of metrics."""
        return len(self._METRICS)

    def __gt__(self, other: object) -> bool:
        """Method implementation."""
        if not isinstance(other, ExcelReportGenerator):
            return NotImplemented
        return len(self) > len(other)

    def __ge__(self, other: object) -> bool:
        """Method implementation."""
        return self == other or self > other

    def __contains__(self, item: str) -> bool:
        """Check if a metric exists."""
        return item.lower() in (name.lower() for name, _ in self._METRICS)
