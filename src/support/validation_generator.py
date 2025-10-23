"""Validation report generators with OOP principles."""

import json
from abc import ABC, abstractmethod
from pathlib import Path
from typing import Any, Union

from src.config.constants import MIN_CONTENT_THRESHOLD

try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    openpyxl = None  # type: ignore
    Font = None  # type: ignore

HAS_OPENPYXL = openpyxl is not None


class BaseValidator(ABC):  # Abstraction
    """Abstract base class for validation report generators."""

    def __init__(self, output_dir: Path) -> None:
        """Initialize validator with output directory."""
        self.__output_dir = output_dir  # Private
        self.__validation_results: dict[str, Any] = {}  # Private

    @property
    def output_dir(self) -> Path:
        """Get output directory."""
        return self.__output_dir

    @property
    def validation_results(self) -> dict[str, Any]:
        """Get validation results."""
        return self.__validation_results.copy()

    @abstractmethod  # Abstraction
    def generate_validation(
        self, toc_data: list[Any], spec_data: list[Any]
    ) -> Path:
        """Generate validation report from TOC and spec data."""

    def _store_results(self, results: dict[str, Any]) -> None:
        """Store validation results."""
        self.__validation_results.update(results)


class XLSValidator(BaseValidator):  # Inheritance
    """Excel-based validation report generator."""

    def generate_validation(
        self, toc_data: list[Any], spec_data: list[Any]
    ) -> Path:  # Polymorphism
        """Generate Excel validation report."""
        if not HAS_OPENPYXL or openpyxl is None:
            raise ImportError("openpyxl required for Excel reports")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"  # type: ignore

        # Create summary
        ws["A1"] = "USB PD Validation Report"  # type: ignore
        ws["A1"].font = Font(bold=True, size=14)  # type: ignore

        status = "PASS" if len(spec_data) > MIN_CONTENT_THRESHOLD else "FAIL"
        metrics: list[tuple[str, Union[int, str]]] = [
            ("TOC Entries", len(toc_data)),
            ("Content Items", len(spec_data)),
            ("Status", status),
        ]

        for i, (metric, value) in enumerate(metrics, 3):
            ws[f"A{i}"] = metric  # type: ignore
            ws[f"B{i}"] = value  # type: ignore

        # Store validation results
        results: dict[str, Any] = {
            "toc_entries": len(toc_data),
            "content_items": len(spec_data),
            "status": status,
        }
        self._store_results(results)

        xlsx_file = self.output_dir / "validation_report.xlsx"
        wb.save(xlsx_file)
        return xlsx_file


class JSONValidator(BaseValidator):  # Inheritance
    """JSON-based validation report generator."""

    def generate_validation(
        self, toc_data: list[Any], spec_data: list[Any]
    ) -> Path:  # Polymorphism
        """Generate JSON validation report."""
        status = "PASS" if len(spec_data) > MIN_CONTENT_THRESHOLD else "FAIL"
        validation_data: dict[str, Any] = {
            "toc_entries": len(toc_data),
            "content_items": len(spec_data),
            "status": status,
            "timestamp": "2024-01-01T00:00:00Z",
        }
        report: dict[str, Any] = {"validation_report": validation_data}

        self._store_results(validation_data)

        json_file = self.output_dir / "validation_report.json"
        with open(json_file, "w", encoding="utf-8") as f:
            json.dump(report, f, indent=2)
        return json_file


class ValidationGeneratorFactory:
    """Factory for validation generators."""

    __VALIDATORS: dict[str, type[BaseValidator]] = {
        "excel": XLSValidator,
        "json": JSONValidator,  # Polymorphism - different implementations
    }

    @classmethod
    def create(cls, validator_type: str, output_dir: Path) -> BaseValidator:
        """Create validator instance."""
        if validator_type not in cls.__VALIDATORS:
            raise ValueError(f"Unknown validator type: {validator_type}")
        return cls.__VALIDATORS[validator_type](output_dir)


def create_validation_report(
    output_dir: Path, toc_file: Path, spec_file: Path
) -> Path:
    """Factory function to create validation report."""
    toc_data: list[Any] = []
    spec_data: list[Any] = []

    # Load TOC data
    try:
        with open(toc_file, encoding="utf-8") as f:
            lines = [line for line in f if line.strip()]
            toc_data = [json.loads(line) for line in lines]
    except (FileNotFoundError, json.JSONDecodeError) as e:
        import logging

        logger = logging.getLogger(__name__)
        logger.debug("Failed to load TOC data: %s", e)
        toc_data = []

    # Load spec data
    try:
        with open(spec_file, encoding="utf-8") as f:
            lines = [line for line in f if line.strip()]
            spec_data = [json.loads(line) for line in lines]
    except (FileNotFoundError, json.JSONDecodeError) as e:
        import logging

        logger = logging.getLogger(__name__)
        logger.debug("Failed to load spec data: %s", e)
        spec_data = []

    validator = ValidationGeneratorFactory.create("excel", output_dir)
    return validator.generate_validation(toc_data, spec_data)
