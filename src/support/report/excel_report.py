"""Excel report generator using openpyxl."""

from pathlib import Path
from typing import Any, Union

from src.config.constants import MIN_CONTENT_THRESHOLD

from .report_generator import BaseReportGenerator  # Import base class

try:
    import openpyxl

    has_openpyxl = True
except ImportError:
    has_openpyxl = False
    openpyxl = None  # type: ignore


class ExcelReportGenerator(BaseReportGenerator):  # Inheritance
    def generate(self, data: dict[str, Any]) -> Path:  # Polymorphism
        if not has_openpyxl or openpyxl is None:
            msg = "openpyxl is required for Excel report generation"
            raise ImportError(msg)

        excel_file = self.output_dir / "validation_report.xlsx"
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
        except Exception as e:
            raise RuntimeError(f"Cannot create Excel workbook: {e}") from e
        try:
            ws.title = "Validation"  # type: ignore

            # Headers
            ws["A1"] = "Metric"  # type: ignore
            ws["B1"] = "Value"  # type: ignore

            # Data
            metrics: list[tuple[str, Union[int, str]]] = [
                ("Pages", data.get("pages", 0)),
                ("Content Items", data.get("content_items", 0)),
                ("TOC Entries", data.get("toc_entries", 0)),
                (
                    "Status",
                    "PASS" if data.get("content_items", 0) > MIN_CONTENT_THRESHOLD else "FAIL",
                ),
            ]

            for i, (metric, value) in enumerate(metrics, 2):
                ws[f"A{i}"] = metric  # type: ignore
                ws[f"B{i}"] = value  # type: ignore
        except Exception as e:
            raise RuntimeError(f"Cannot write Excel data: {e}") from e

        try:
            wb.save(excel_file)
        except OSError as e:
            raise RuntimeError(f"Cannot save Excel file: {e}") from e
        return excel_file
